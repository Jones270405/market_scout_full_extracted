# comparison_report_agent/agent.py
"""
Comparison Report Sub-Agent
Builds side-by-side competitor comparison tables and Excel workbooks.
"""

import os
from datetime import datetime
from google.adk.agents import LlmAgent
from google.adk.models.lite_llm import LiteLlm
from google.adk.tools import FunctionTool
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference

OUTPUT_DIR = os.environ.get("MARKET_SCOUT_OUTPUT_DIR", os.path.join(os.getcwd(), "outputs"))
os.makedirs(OUTPUT_DIR, exist_ok=True)
EXCEL_FILE = os.path.join(OUTPUT_DIR, "market_scout_data.xlsx")


def update_excel(all_runs: list) -> str:
    """
    Writes all historical runs to a persistent Excel workbook with two sheets:
      - All Features : every feature row colour-coded by status
      - Summary      : per-run counts with a bar chart
    Returns the Excel file path.
    """
    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "All Features"

    headers     = ["Run Date", "Company", "Feature", "Category", "Date", "Status", "URL"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell           = ws1.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    color_map = {
        "WEEK"      : "C6EFCE",
        "MONTH"     : "FFEB9C",
        "YEAR"      : "DDEBF7",
        "UNVERIFIED": "F2F2F2",
        "STALE"     : "FFC7CE",
    }

    row = 2
    for run in all_runs:
        for f in run.get("features", []):
            color = color_map.get(f.get("status", ""), "FFFFFF")
            fill  = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws1.cell(row=row, column=1, value=run.get("run_date", ""))
            ws1.cell(row=row, column=2, value=run.get("company", ""))
            ws1.cell(row=row, column=3, value=f.get("feature", ""))
            ws1.cell(row=row, column=4, value=f.get("category", ""))
            ws1.cell(row=row, column=5, value=f.get("date", "unknown"))
            ws1.cell(row=row, column=6, value=f.get("status", ""))
            ws1.cell(row=row, column=7, value=f.get("url", ""))
            for col in range(1, 8):
                ws1.cell(row=row, column=col).fill      = fill
                ws1.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    col_widths = [15, 15, 45, 15, 12, 12, 50]
    for col, width in enumerate(col_widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Company", "Run Date", "Total", "Week", "Month", "Year"])
    ws2["A1"].font = Font(bold=True)

    for run in all_runs:
        features = run.get("features", [])
        ws2.append([
            run.get("company", ""),
            run.get("run_date", ""),
            len(features),
            sum(1 for f in features if f.get("status") == "WEEK"),
            sum(1 for f in features if f.get("status") in ["WEEK", "MONTH"]),
            sum(1 for f in features if f.get("status") in ["WEEK", "MONTH", "YEAR"]),
        ])

    if all_runs:
        chart              = BarChart()
        chart.type         = "col"
        chart.title        = "Features by Company"
        chart.y_axis.title = "Features"
        data = Reference(ws2, min_col=3, max_col=6, min_row=1, max_row=len(all_runs) + 1)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=len(all_runs) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws2.add_chart(chart, "H2")

    wb.save(EXCEL_FILE)
    return EXCEL_FILE


def build_comparison_table(runs: list) -> str:
    """
    Builds a markdown side-by-side comparison table for multiple companies.
    Each run dict must contain: company, summary (total/week/month/year/unver).
    """
    if not runs:
        return "_No runs provided for comparison._"

    header = "| Metric |"
    sep    = "|--------|"
    rows   = {
        "Total Features": "| Total Features |",
        "Last 7 Days"   : "| Last 7 Days 🟢 |",
        "Last 30 Days"  : "| Last 30 Days 🟡 |",
        "Last 365 Days" : "| Last 365 Days 🔵 |",
        "Unverified"    : "| Unverified ⚪ |",
    }

    for run in runs:
        company = run.get("company", "?")
        summary = run.get("summary", {})
        header += f" {company} |"
        sep    += "----------|"
        rows["Total Features"] += f" {summary.get('total', 0)} |"
        rows["Last 7 Days"]    += f" {summary.get('week', 0)} |"
        rows["Last 30 Days"]   += f" {summary.get('month', 0)} |"
        rows["Last 365 Days"]  += f" {summary.get('year', 0)} |"
        rows["Unverified"]     += f" {summary.get('unver', 0)} |"

    table  = header + "\n" + sep + "\n"
    table += "\n".join(rows.values())
    return table


excel_tool      = FunctionTool(func=update_excel)
comparison_tool = FunctionTool(func=build_comparison_table)

comparison_report_agent = LlmAgent(
    name="comparison_report_agent",
    model=LiteLlm(model="groq/llama-3.1-8b-instant"),
    description="Builds Excel workbooks and comparison tables for multi-company runs.",
    instruction=(
        "You are a Comparison Report Agent. "
        "When given a list of run dicts: "
        "1. Call update_excel(all_runs) to persist data to Excel. "
        "2. Call build_comparison_table(runs) to produce a markdown comparison table. "
        "Return both outputs exactly as received. Do not summarise or modify them."
    ),
    tools=[excel_tool, comparison_tool],
)
